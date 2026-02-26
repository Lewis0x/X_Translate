from __future__ import annotations

import json
import re
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import asdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from typing import List

from doc_translator.translator import TranslationConfig
from doc_translator.translator import create_translator

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


@dataclass
class CompareResult:
    name: str
    provider: str
    model: str
    score: float
    success_ratio: float
    number_match_ratio: float
    length_stability: float
    sample_size: int
    error: str = ""


def collect_sample_texts(files: Iterable[Path], sample_size: int) -> List[str]:
    rows: List[str] = []
    for file in files:
        suffix = file.suffix.lower()
        if suffix == ".docx":
            rows.extend(_extract_docx(file))
        elif suffix == ".xlsx":
            rows.extend(_extract_xlsx(file))
        elif suffix == ".pdf":
            rows.extend(_extract_pdf(file))
        if len(rows) >= sample_size:
            break
    return rows[:sample_size]


def choose_best_profile(
    profiles: list[tuple[str, TranslationConfig]],
    sample_texts: List[str],
    logger,
    compare_report_file: Path,
) -> tuple[str, TranslationConfig, List[CompareResult]]:
    results: List[CompareResult] = []

    for name, config in profiles:
        logger.info("开始对比模型: %s (%s/%s)", name, config.provider, config.model)
        try:
            translator = create_translator(config)
            translated = translator.translate(sample_texts)
            score, success_ratio, number_match_ratio, length_stability = _score(sample_texts, translated)
            results.append(
                CompareResult(
                    name=name,
                    provider=config.provider,
                    model=config.model,
                    score=score,
                    success_ratio=success_ratio,
                    number_match_ratio=number_match_ratio,
                    length_stability=length_stability,
                    sample_size=len(sample_texts),
                )
            )
            logger.info("模型对比分数: %s -> %.4f", name, score)
        except Exception as exc:
            results.append(
                CompareResult(
                    name=name,
                    provider=config.provider,
                    model=config.model,
                    score=0.0,
                    success_ratio=0.0,
                    number_match_ratio=0.0,
                    length_stability=0.0,
                    sample_size=len(sample_texts),
                    error=str(exc),
                )
            )
            logger.warning("模型对比失败: %s -> %s", name, exc)

    if not results:
        raise RuntimeError("没有可用的模型对比结果")

    sorted_results = sorted(results, key=lambda item: item.score, reverse=True)
    winner_name = sorted_results[0].name
    winner_config = next(cfg for name, cfg in profiles if name == winner_name)

    compare_report_file.parent.mkdir(parents=True, exist_ok=True)
    compare_report_file.write_text(
        json.dumps([asdict(item) for item in sorted_results], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return winner_name, winner_config, sorted_results


def _score(source: List[str], target: List[str]) -> tuple[float, float, float, float]:
    total = max(1, len(source))
    pairs = list(zip(source, target, strict=False))
    if not pairs:
        return 0.0, 0.0, 0.0, 0.0

    success_count = sum(1 for _src, tgt in pairs if bool(str(tgt).strip()))
    number_match_count = sum(1 for src, tgt in pairs if _numbers(src) == _numbers(tgt))
    length_scores = [_length_score(src, tgt) for src, tgt in pairs]

    success_ratio = success_count / total
    number_match_ratio = number_match_count / total
    length_stability = sum(length_scores) / max(1, len(length_scores))
    score = success_ratio * 0.5 + number_match_ratio * 0.3 + length_stability * 0.2
    return score, success_ratio, number_match_ratio, length_stability


def _numbers(text: str) -> list[str]:
    return re.findall(r"\d+(?:[\.,]\d+)?", str(text))


def _length_score(src: str, tgt: str) -> float:
    source_len = max(1, len(str(src).strip()))
    target_len = len(str(tgt).strip())
    delta = abs(target_len - source_len) / source_len
    return max(0.0, 1.0 - delta)


def _extract_docx(path: Path) -> List[str]:
    texts: List[str] = []
    with zipfile.ZipFile(path, "r") as zf:
        names = [name for name in zf.namelist() if name.startswith("word/") and name.endswith(".xml")]
        for name in names:
            if not _is_translatable_docx_part(name):
                continue
            root = ET.fromstring(zf.read(name))
            for node in root.iter(f"{{{W_NS}}}t"):
                text = (node.text or "").strip()
                if text:
                    texts.append(text)
    return texts


def _extract_xlsx(path: Path) -> List[str]:
    from openpyxl import load_workbook

    texts: List[str] = []
    wb = load_workbook(path)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() and not cell.value.startswith("="):
                    texts.append(cell.value)
    return texts


def _extract_pdf(path: Path) -> List[str]:
    import fitz

    texts: List[str] = []
    doc = fitz.open(path)
    for page in doc:
        for block in page.get_text("blocks"):
            _x0, _y0, _x1, _y1, text, _block_no, block_type = block
            if int(block_type) != 0:
                continue
            raw = (text or "").strip()
            if raw:
                texts.append(raw)
    doc.close()
    return texts


def _is_translatable_docx_part(name: str) -> bool:
    if name == "word/document.xml":
        return True
    if name.startswith("word/header") and name.endswith(".xml"):
        return True
    if name.startswith("word/footer") and name.endswith(".xml"):
        return True
    if name in {"word/footnotes.xml", "word/endnotes.xml", "word/comments.xml"}:
        return True
    return False
