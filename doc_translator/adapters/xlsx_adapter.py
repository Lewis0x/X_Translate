from __future__ import annotations

from collections.abc import Callable
from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook

from doc_translator.adapters.base import ProcessStats
from doc_translator.glossary import Glossary
from doc_translator.translator import TranslatorProtocol


class XlsxAdapter:
    suffixes = (".xlsx",)

    def process(
        self,
        input_file: Path,
        output_file: Path,
        translator: TranslatorProtocol,
        glossary: Glossary,
        progress_callback: Callable[[int, int], None] | None = None,
    ) -> ProcessStats:
        output_file.parent.mkdir(parents=True, exist_ok=True)

        workbook = load_workbook(input_file)
        stats = ProcessStats()

        cells: List[Tuple[object, dict[str, str]]] = []
        source_texts: List[str] = []

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    if value.startswith("="):
                        continue
                    if not value.strip():
                        continue
                    prepared, placeholders, lock_hits = glossary.preprocess_locks(value)
                    cells.append((cell, placeholders))
                    source_texts.append(prepared)
                    stats.glossary_hits += lock_hits

        translated = translator.translate(source_texts, progress_callback=progress_callback)
        for index, translated_text in enumerate(translated):
            restored, hits = glossary.postprocess(translated_text, cells[index][1])
            cells[index][0].value = restored
            stats.glossary_hits += hits

        stats.segments_total = len(source_texts)
        stats.segments_translated = len(translated)

        workbook.save(output_file)
        return stats
